import os
import ollama
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# GENERATE_SUMMERY_FOR = ['SPA3_CSWV3.1.1_JBW41R XLP11 ', 'SPA3_CSWV3.0.1_JBW41R XLP1109', 'SPA3_INT-3873 (XLT1026)', 'SPA3_INT-3816 (XLT1051) ', 'SPA3_INT-3678 (XLT1051)', 'SPA3_INT-3743 (XLT1026)']
# GENERATE_SUMMERY_FOR = ['SPA3_CSWV3.1.1_JBW41R XLP11 ', 'SPA3_CSWV3.0.1_JBW41R XLP1109']
GENERATE_SUMMERY_FOR = ["all"]
GENERATE_SUMMERY_FOR = []

NO_GENERATE_DEBUG = False

EXTRACTION_MAP_1 = {
    "SW": "A3",
    "Vehicle ID": "A4",
    "KPI": "H1",
}

SW_ID_LEN = 5 # use last 5 chars as SW identifier

COLUMNS = ["Test Name", "Test week"] + \
    list(EXTRACTION_MAP_1.keys()) + \
    ["Severity 5: Critical",
    "Severity 4: Major",
    "Severity 3: Moderate","Summary"]

DEP_TEST_SHEET_NAME = "SPA3_"
LOOP1_IDX = 10 # column K (index 10) is where probability values start in the sheet
MAX_LOOPS = 35
MAX_LOOP_IDX = LOOP1_IDX + MAX_LOOPS
SUMMARY_COL_IDX = 5  # column F (zero-based index in values_only row tuple)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

SOURCE_FILE = os.path.join(BASE_DIR, "data", "PKV Dependability SPA3 Test workbook.xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, "data", "PKV Dependability Overview.xlsx")
DEBUG_TEMP_FILE_BASE = os.path.join(BASE_DIR, "data", "debug_temp_")

SEVERITY_LABELS = [
    "Severity 5: Critical",
    "Severity 4: Major",
    "Severity 3: Moderate",
    "Severity 2: Minor",
    "Severity 1: Low",
    "Analysis",
    "New",
]

SEVERITY_LABELS_OF_INTEREST = [
    "Severity 5: Critical",
    "Severity 4: Major",
    "Severity 3: Moderate"
]

LLM_MODEL = "phi3:mini"
LLM_OPTIONS = {
    "temperature": 0.6,
    "num_predict": 500,
    "num_ctx": 8192
}
LLM_TAG = " (AUTO GENERATED | LLM model: " + LLM_MODEL + ", options: " + str(LLM_OPTIONS) + ")"

SUMMARY_LENGTH = "200" # desired summary length in words, span "x-y" format is ok

def to_numeric_score(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        normalized = value.strip()
        if not normalized:
            return None
        if str(normalized).lower() in {"ok", "nt"}:
            return 0.0
        # normalized = normalized.replace(",", ".")
        try:
            return float(normalized)
        except ValueError:
            return None
    return None


def calc_row_probability(row):
    # Count probability entries
    values = []
    #TODO: max col should be determined based on first None in VIRA task row
    for cell in row[LOOP1_IDX:MAX_LOOP_IDX]: #start from column K (index 10)
        numeric_value = to_numeric_score(cell)
        if numeric_value is not None:
            values.append(numeric_value)

    # prob = sum(values) / len(values) if values else None
    return values


def count_arts_per_severity(ws):
    """Count ART entries and collect per-issue details under each section."""
    counts = {}
    probability = {}
    details = {}
    current_label = None

    for row in ws.iter_rows(min_col=1, max_col=ws.max_column, values_only=True):
        cell_value = row[0]
        if cell_value in SEVERITY_LABELS:
            current_label = cell_value
            counts[current_label] = 0
            probability[current_label] = []
            details[current_label] = []

        elif current_label is not None and cell_value is not None and isinstance(cell_value, str) and cell_value != '.':

            counts[current_label] += 1
            prob_list = calc_row_probability(row)
            probability[current_label] += prob_list
            summary_value = row[SUMMARY_COL_IDX]
            details[current_label].append({
                "issue": cell_value,
                "summary": summary_value,
                "probs": prob_list,
            })
    
    # print(details)
        
    return counts, probability, details

def create_summary_prompt(details, sheet_name):
    prompt = f"You are a vehicle reliability analyst. Below I present a number of fault reports divided into {len(SEVERITY_LABELS_OF_INTEREST)} severity levels, indicating how critical the fault/issue is. Each fault report contains an \"ID\" for identification, a short \"Summary\" and a \"Frequency\" in the range 0-5, where values close to 5 indicates that the issue appers more frequently. The faults are seen on a specific singular vehicle running an specific software, the reason for the summary is to indicate the overall reliability of the software. Now you will summarize the fault landscape, you will put the majority of the focus on the faults with high severity and frequency. Your summary should be {SUMMARY_LENGTH} words long, not longer. Below are all of the fault reports:\n\n"
    for severity, issues in details.items():
        if severity not in SEVERITY_LABELS_OF_INTEREST:
            continue
        prompt += f"{severity}\n{'-'*80} \n"
        for issue in issues:
            mean_prob = float(round(sum(issue['probs'])/len(issue['probs']),3))
            if mean_prob > 0.0: # remove 0 prob issues
                prompt += f"- ID: {issue['issue']}\n"
                prompt += f"  Summary: {issue['summary']}\n"
                prompt += f"  Frequency (0-5): {mean_prob}\n\n"
        prompt += f"{'-'*80}\n\n"
    # prompt += f"Summarize the above information in a concise way, focusing on the most severe and frequent issues. Provide insights on potential root causes and suggest areas for further investigation.\nSummary:\n"
    prompt += f"Summarize the above information in a concise way, focusing on the most severe and frequent issues, aswell as less sever issues if they have very high frequency. You do not need to mention low severity, low frequency issues if there is no direct reason to do so. Do not include the \"Frequency\" values in the summary, rather use word, e.g. high frequency. Keep the summary factual and objective and very shortly mentioned the overall \"grade\" of the tested vehicle software combination.\nYour {SUMMARY_LENGTH}-word summary:\n"

    with open(DEBUG_TEMP_FILE_BASE+sheet_name+".txt", "w", encoding="utf-8") as f:
        f.write(f"################ Sheet: {sheet_name} ################\n")
        f.write(f"Prompt ({len(prompt.split())} words):\n")
        f.write(prompt)
        f.write("\n")

    return prompt

def generate_summary(details, sheet_name):
    prompt = create_summary_prompt(details, sheet_name)
    if not NO_GENERATE_DEBUG:
        summary = ollama.generate(model=LLM_MODEL, prompt=prompt, options=LLM_OPTIONS)
    else:
        summary = type('obj', (object,), {'response' : "SUM"})()
    # summary="hej"
    # print("Summary response:", response)
    with open(DEBUG_TEMP_FILE_BASE+sheet_name+".txt", "a", encoding="utf-8") as f:
        f.write(summary.response + LLM_TAG)
        f.write("\n")

    return summary.response + LLM_TAG

def remove_prefix(value, prefix):
    if isinstance(value, str) and value.startswith(prefix):
        return value[len(prefix):]
    return value

wb = load_workbook(SOURCE_FILE, data_only=True)

print("Sheet names:", wb.sheetnames)

# Load Test week lookup from Länkar sheet
df_lankar = pd.read_excel(SOURCE_FILE, sheet_name="Länkar", header=0)
df_lankar["SW label"] = df_lankar["SW label"].astype("string").str[-SW_ID_LEN:]
sw_to_test_week = dict(zip(df_lankar["SW label"].to_list(), df_lankar["Test week"]))

# print(df_lankar["SW label"])

df_out = pd.DataFrame(columns=COLUMNS)

for sheet_name in wb.sheetnames:
    if sheet_name[0:len(DEP_TEST_SHEET_NAME)] == DEP_TEST_SHEET_NAME:
        ws = wb[sheet_name]
        df_out.at[sheet_name, "Test Name"] = sheet_name
        print(f"Processing sheet: {sheet_name}")
        for key, cell in EXTRACTION_MAP_1.items():
            value = ws[cell].value
            if key == "SW":
                value = remove_prefix(value, "SW under test: ")
            if key == "Vehicle ID":
                value = remove_prefix(value, "Vehicle ID: ")
                value = re.sub(r'\s+', '', value).split('/', 1)[-1] # remove reg nr. if included
            if key == "KPI" and value is not None:
                value = str(value)
            # print(f"{key} ({cell}): {value}")
            df_out.at[sheet_name, key] = value

        # print(df_out.at[sheet_name, "SW"][-SW_ID_LEN:],":",sw_to_test_week.get(df_out.at[sheet_name, "SW"][-SW_ID_LEN:]))
        df_out.at[sheet_name, "Test week"] = sw_to_test_week.get(df_out.at[sheet_name, "SW"][-SW_ID_LEN:])

        # Count issues per severity level
        counts, probability, details = count_arts_per_severity(ws)
        for label in SEVERITY_LABELS:
            if label in df_out.columns:
                prob_list = probability.get(label)
                # print(label, "prob list:", prob_list)
                avg_prob = sum(prob_list) / len(prob_list) if prob_list else 0
                # print(prob_list)
                # print(avg_prob)
                df_out.at[sheet_name, label] = "Count: " + str(counts.get(label, 0)) + "\t\t (avg prob: " + str(round(avg_prob,1)) + ")"
        
        if ("all" in GENERATE_SUMMERY_FOR) or (sheet_name in GENERATE_SUMMERY_FOR):
            print("Start summary generation for sheet:", sheet_name)
            df_out.at[sheet_name, "Summary"] = generate_summary(details, sheet_name)
        else:
            print("Skipped summary generation for sheet:", sheet_name)
            df_out.at[sheet_name, "Summary"] = None # use None to indicate not generated, needed for filling only regenerated summary cells

# for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
#     print(row)

# print(ws[15].value)

def create_excel(df):
    df.to_excel(OUTPUT_FILE, sheet_name="Overview", index=False)

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-fit columns
    for column in ws.columns:
        ws.column_dimensions[column[0].column_letter].width = 30

    # Summary column (I): make wider and increase row height for long text cells
    ws.column_dimensions["I"].width = 70
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 90

    # Align all populated cells to top-left
    top_left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = top_left_alignment

    ws.auto_filter.ref = ws.dimensions

    wb.save(OUTPUT_FILE)
    wb.close()

def update_excel(df):
    """Update existing Excel file with new data while preserving formatting and layout."""
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    # Get the summary column index (last column in df, 1-based for openpyxl)
    summary_col_idx = len(df.columns)
    
    # Write new data starting from row 2
    for r_idx, row in enumerate(df.values, start=2):
        for c_idx, value in enumerate(row, start=1):
            # For Summary column, only write if value is not None/empty
            if c_idx == summary_col_idx:
                # Check for None or pandas NaN
                if value is None or (isinstance(value, float) and pd.isna(value)):
                    continue  # Skip, keep existing value
            ws.cell(row=r_idx, column=c_idx).value = value

    # # Re-apply row heights for data rows to accommodate content
    # for row_idx in range(2, ws.max_row + 1):
    #     ws.row_dimensions[row_idx].height = 90

    wb.save(OUTPUT_FILE)
    wb.close()


# print(df_out["Test Name"].tolist())

# Check if output file exists and is valid; create new or update existing
if os.path.exists(OUTPUT_FILE):
    try:
        print(f"Updating existing file: {OUTPUT_FILE}")
        update_excel(df_out)
    except Exception as e:
        print(f"Warning: Could not update existing file ({e}). Creating new file instead.")
        create_excel(df_out)
else:
    print(f"Creating new file: {OUTPUT_FILE}")
    create_excel(df_out)

os.startfile(OUTPUT_FILE)



# def excel_to_py_index(cell_reference):
#     row_index, column_index = coordinate_to_tuple(cell_reference)
#     return row_index-2, column_index-1

# print("A3:", excel_to_py_index("A3"))  # Should print (2, 0)

# sheets = pd.read_excel(SOURCE_FILE, sheet_name=None)

# print(sheets.keys())

# for sheet_name, df in sheets.items():
#     if sheet_name[0:len(DEP_TEST_SHEET_NAME)] == DEP_TEST_SHEET_NAME:
#         print(f"Processing sheet: {sheet_name}")
#         print(df.iat[excel_to_py_index("F8")])
#         for key, cell in EXTRACTION_MAP_1.items():
#             row_index, column_index = excel_to_py_index(cell)
#             value = df.iat[row_index, column_index]
#             print(f"{key} ({cell}): {value}")